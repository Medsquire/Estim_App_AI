#!/usr/bin/env python3
"""
Convert Excel SOR file to JSON format - Improved version
Handles complex Excel structure with merged cells and multi-row items
"""
import json
import re
import os
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("openpyxl not installed. Attempting to install...")
    os.system("pip install openpyxl")
    import openpyxl


def is_valid_item_number(value):
    """Check if value is a valid item number"""
    if not value:
        return False
    
    str_val = str(value).strip()
    
    # Skip common headers
    if str_val.lower() in ['item no', 'item no.', 'sno', '#', 'description', 'unit', 'rate', 'qty', 'chapter', 'sl. no.', 'item \nno.']:
        return False
    
    # Skip section headers
    if str_val.lower() in ['supply portion', 'labour portion', 'a', 'b', 'description of the item']:
        return False
    
    # Valid items start with - or number
    if str_val.startswith('-') or re.match(r'^\d+', str_val):
        return True
    
    return False


def parse_rate(value):
    """Extract numeric rate from value - ensure positive value"""
    if not value:
        return 0
    
    try:
        # If it's already a number
        if isinstance(value, (int, float)):
            rate = float(value)
            return abs(rate)  # Ensure positive
        
        # Extract number from string
        str_val = str(value).strip()
        # Remove common non-numeric characters but keep digit and decimal
        cleaned = re.sub(r'[^\d.]', '', str_val)
        if cleaned:
            rate = float(cleaned)
            return abs(rate)  # Ensure positive
    except (ValueError, TypeError):
        pass
    
    return 0


def convert_excel_to_json_improved():
    """Convert SOR Excel file to JSON format with improved parsing"""
    
    # Define paths
    script_dir = Path(__file__).parent.parent.parent
    excel_file = script_dir / "uploads" / "SOR_files" / "181025... Corrected Final Copy SCR SOR 2024 Supply & Labour (1).xlsx"
    output_file = script_dir / "Json" / "SOR" / "SOR_EXCEL.json"
    
    # Check if Excel file exists
    if not excel_file.exists():
        print(f"Error: Excel file not found at {excel_file}")
        return
    
    try:
        # Load the workbook
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        
        print(f"Reading from sheet: {ws.title}")
        print(f"Dimensions: {ws.dimensions}")
        
        items = []
        current_chapter = 1
        current_item = None
        item_count = 0
        
        # Get all rows with values
        rows_data = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            if any(cell for cell in row):
                rows_data.append(row)
        
        print(f"Total non-empty rows: {len(rows_data)}")
        
        # Process rows
        for row_idx, row in enumerate(rows_data):
            if not row or not any(row):
                continue
            
            col1 = str(row[0]).strip() if row[0] else ""
            col2 = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            col3 = str(row[2]).strip() if len(row) > 2 and row[2] else ""
            col4 = str(row[3]).strip() if len(row) > 3 and row[3] else ""
            col5 = str(row[4]).strip() if len(row) > 4 and row[4] else ""
            
            # Skip empty rows
            if not col1 and not col2:
                continue
            
            # Detect chapter changes
            if "chapter" in col1.lower():
                try:
                    match = re.search(r'\d+', col1)
                    if match:
                        current_chapter = int(match.group())
                        print(f"  Found Chapter {current_chapter}")
                except:
                    pass
                continue
            
            # Check if this row contains a valid item
            if is_valid_item_number(col1):
                # Check if we have a complete item (should have item_no, description, unit, rate)
                # Item numbers might be like "101", "101 (a)", etc.
                item_no = col1
                description = col2
                unit = col3
                
                # Try to get rate - it might be in col4 or col5
                rate_val = parse_rate(col4) or parse_rate(col5)
                
                # Skip if no description
                if not description or description.lower() in ['description', 'description of the item']:
                    continue
                
                # Create item
                item = {
                    "item_no": item_no,
                    "description": description,
                    "unit": unit if unit else "No.",
                    "rate": rate_val,
                    "chapter": current_chapter,
                    "page_no": row_idx + 1,
                    "fileType": "SOR",
                    "reference_id": "SOR File/2024-2025"
                }
                items.append(item)
                item_count += 1
                
                # Print first 10 items for verification
                if item_count <= 10:
                    print(f"  Item {item_count}: {item_no} | {description[:60]}... | Unit: {unit} | Rate: {rate_val}")
        
        print(f"\nTotal items extracted: {item_count}")
        
        # Create output structure
        output_data = {
            "document": "South Central Railway Signal & Telecommunication Schedule of Rates 2024",
            "source": "181025... Corrected Final Copy SCR SOR 2024 Supply & Labour (1).xlsx",
            "conversion_date": "2026-06-21",
            "total_items": len(items),
            "items": items
        }
        
        # Write to JSON file
        output_file.parent.mkdir(parents=True, exist_ok=True)
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(output_data, f, indent=2, ensure_ascii=False)
        
        print(f"\n✓ Successfully converted {len(items)} items to JSON")
        print(f"✓ Output saved to: {output_file}")
        
        # Show statistics
        if items:
            print(f"\nSample items:")
            for i, item in enumerate(items[:3]):
                print(f"\nItem {i+1}:")
                print(json.dumps(item, indent=2))
        
    except Exception as e:
        print(f"Error during conversion: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    convert_excel_to_json_improved()
